import csv
import unicodedata
from io import StringIO, BytesIO
from collections import Counter
from datetime import datetime
from fastapi import UploadFile
from openpyxl import load_workbook


def decode_bytes_to_text(raw_bytes: bytes) -> str:
    for encoding_name in ("utf-8-sig", "utf-16", "latin1"):
        try:
            return raw_bytes.decode(encoding_name)
        except UnicodeDecodeError:
            continue
    return raw_bytes.decode("utf-8", errors="ignore")


def dict_reader_from_text(text: str) -> csv.DictReader:
    return csv.DictReader(StringIO(text, newline=""), delimiter=";")


def normalize_key(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text or "").encode("ascii", "ignore").decode()
    return " ".join(normalized.lower().split())


def parse_number(raw_value) -> float | None:
    if raw_value is None:
        return None
    if isinstance(raw_value, (int, float)):
        return float(raw_value)
    cleaned_value = str(raw_value).strip()
    if not cleaned_value:
        return None
    cleaned_value = cleaned_value.replace("R$", "").replace(" ", "")
    cleaned_value = cleaned_value.replace(".", "", cleaned_value.count(".") - 1)
    cleaned_value = cleaned_value.replace(",", ".")
    try:
        return float(cleaned_value)
    except ValueError:
        return None


def format_expense(amount: float) -> str:
    return f"-{int(amount)}" if float(amount).is_integer() else f"-{amount:.2f}"


def load_invoice_transactions(file: UploadFile) -> list[dict]:
    raw_bytes = file.file.read()
    text = decode_bytes_to_text(raw_bytes)
    file.file.seek(0)
    reader = dict_reader_from_text(text)
    header_map = {normalize_key(h): h for h in (reader.fieldnames or [])}
    date_key = header_map.get("data")
    description_key = header_map.get("estabelecimento")
    value_key = header_map.get("valor")
    result = []
    for row in reader:
        date_value = row.get(date_key, "").strip() if date_key else ""
        description_value = row.get(description_key, "").strip() if description_key else ""
        numeric_value = parse_number(row.get(value_key)) if value_key else None
        if not description_value or numeric_value is None:
            continue
        if numeric_value < 0:
            continue
        expense_value = format_expense(numeric_value)
        result.append({"date": date_value, "description": description_value, "value": expense_value})
    return result


def load_computed_transactions_csv(file: UploadFile) -> list[dict]:
    raw_bytes = file.file.read()
    text = decode_bytes_to_text(raw_bytes)
    file.file.seek(0)
    reader = dict_reader_from_text(text)
    header_map = {normalize_key(h): h for h in (reader.fieldnames or [])}
    date_key = header_map.get("date")
    description_key = header_map.get("description")
    value_key = header_map.get("value")
    result = []
    for row in reader:
        date_value = row.get(date_key, "").strip() if date_key else ""
        description_value = row.get(description_key, "").strip() if description_key else ""
        numeric_value = parse_number(row.get(value_key)) if value_key else None
        if not description_value or numeric_value is None:
            continue
        expense_value = format_expense(abs(numeric_value))
        result.append({"date": date_value, "description": description_value, "value": expense_value})
    return result


def load_computed_transactions_xlsx(file: UploadFile) -> list[dict]:
    raw_bytes = file.file.read()
    file.file.seek(0)
    wb = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {normalize_key(h): i for i, h in enumerate(headers)}
    i_date = idx.get("date")
    i_desc = idx.get("description")
    i_val = idx.get("value")
    result = []
    for r in rows:
        date_cell = r[i_date] if i_date is not None and i_date < len(r) else ""
        desc_cell = r[i_desc] if i_desc is not None and i_desc < len(r) else ""
        value_cell = r[i_val] if i_val is not None and i_val < len(r) else None
        date_value = date_cell.strftime("%d/%m/%Y") if isinstance(date_cell, datetime) else str(date_cell or "").strip()
        description_value = str(desc_cell or "").strip()
        numeric_value = parse_number(value_cell)
        if not description_value or numeric_value is None:
            continue
        expense_value = format_expense(abs(numeric_value))
        result.append({"date": date_value, "description": description_value, "value": expense_value})
    return result


def load_computed_transactions(file: UploadFile) -> list[dict]:
    name = (file.filename or "").lower()
    ctype = (file.content_type or "").lower()
    if name.endswith(".xlsx") or "excel" in ctype or "spreadsheetml" in ctype:
        return load_computed_transactions_xlsx(file)
    return load_computed_transactions_csv(file)


def filter_transactions(invoice_file: UploadFile, computed_file: UploadFile | None):
    invoice_transactions = load_invoice_transactions(invoice_file)
    if not computed_file:
        return sorted(invoice_transactions, key=lambda x: x["date"])
    computed_transactions = load_computed_transactions(computed_file)
    computed_counts = Counter((t["description"], t["value"]) for t in computed_transactions)
    consumed_counts = Counter()
    filtered = []
    for tx in sorted(invoice_transactions, key=lambda x: x["date"]):
        key = (tx["description"], tx["value"])
        if consumed_counts[key] < computed_counts[key]:
            consumed_counts[key] += 1
            continue
        filtered.append(tx)
    return filtered


def generate_csv(transactions: list[dict]) -> StringIO:
    output = StringIO()
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_ALL)
    writer.writerow(["Data", "Descrição", "Valor", "Conta", "Categoria"])
    for tx in transactions:
        writer.writerow([tx["date"], tx["description"], tx["value"], "Carteira", "Alimentação"])
    output.seek(0)
    return output
